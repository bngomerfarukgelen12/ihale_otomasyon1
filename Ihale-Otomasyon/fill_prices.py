import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

# Tüm birim fiyatları oku
fiyat_df = pd.read_excel(r'c:\Users\argge\OneDrive\Masaüstü\Ihale-Otomasyon\excel\csb_tum_fiyatlar.xlsx')
fiyat_df['Poz No'] = fiyat_df['Poz No'].astype(str)

# Poz No -> Birim Fiyat şeklinde dictionary yap
fiyat_dict = dict(zip(fiyat_df['Poz No'], fiyat_df['Birim Fiyat (TL)']))
print(f"Birim fiyat sözlüğü oluşturuldu: {len(fiyat_dict)} poz")

# Belediye cetvelini aç
cetveli_path = r'c:\Users\argge\OneDrive\Masaüstü\Ihale-Otomasyon\excel\belediye teklif cetveli.xlsx'
wb = openpyxl.load_workbook(cetveli_path)
ws = wb.active

# Veri satırları: 6'dan 1883'e (1878 poz)
start_row = 6
end_row = 1883

updated_count = 0
missing_count = 0

print("\nBirim fiyatları güncelleniyor...")
for row in range(start_row, end_row):
    # B sütunundan Poz No'yu al
    poz_no = ws[f'B{row}'].value
    
    if poz_no is None:
        continue
    
    # F sütununa (Teklif Edilen Birim Fiyat) birim fiyat yaz
    birim_fiyat = fiyat_dict.get(str(poz_no))
    
    if birim_fiyat is not None:
        ws[f'F{row}'].value = birim_fiyat
        ws[f'F{row}'].number_format = '#,##0.00'
        ws[f'F{row}'].alignment = Alignment(horizontal='right')
        updated_count += 1
    else:
        missing_count += 1
    
    # G sütununa (Tutarı) formül yaz: = Miktarı (E) * Birim Fiyat (F)
    if ws[f'E{row}'].value is not None:
        ws[f'G{row}'].value = f'=E{row}*F{row}'
        ws[f'G{row}'].number_format = '#,##0.00'
        ws[f'G{row}'].alignment = Alignment(horizontal='right')
    
    if row % 200 == 0:
        print(f"  Satır {row} işlendi...")

print(f"\nBirim fiyatları güncellendi: {updated_count} poz")
print(f"Birim fiyat bulunamadı: {missing_count} poz")

# Toplam formülü kontrol et
total_row = end_row
if ws[f'G{total_row}'].value is None:
    ws[f'G{total_row}'].value = f'=SUM(G{start_row}:G{end_row-1})'
    ws[f'G{total_row}'].number_format = '#,##0.00'
    ws[f'G{total_row}'].alignment = Alignment(horizontal='right')
    print(f"\nToplam Tutar formülü eklendi: {ws[f'G{total_row}'].value}")

# Dosyayı kaydet
wb.save(cetveli_path)
print(f"\nDosya kaydedildi: {cetveli_path}")
