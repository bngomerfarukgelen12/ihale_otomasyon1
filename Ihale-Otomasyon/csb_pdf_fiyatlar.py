"""CSB PDF birim fiyat listelerinden poz fiyatlarını çıkarır."""
from __future__ import annotations

import logging
import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

import fitz
import openpyxl

from logger_config import setup_logger

logger = setup_logger("ihale_otomasyonu")

ROOT = Path(__file__).parent
PDF_DIR = ROOT / "pdfler"
CACHE_XLSX = ROOT / "excel" / "csb_tum_fiyatlar.xlsx"
CACHE_TANIM_XLSX = ROOT / "excel" / "csb_tum_tanimlar.xlsx"

POZ_RE = re.compile(r"\d{2}\.\d{3}\.\d{4}")
FIYAT_RE = re.compile(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}")
BIRIM_RE = re.compile(
    r"\b(?:\d+\s+)?(?:1000\s+Ad|100\s+m[²2³3]?|m[²2³3]?|metrekare|metreküp|metre|Ton|Ad|kg|lt|g|t|ad)\b\.?",
    re.I,
)
TR_MAP = str.maketrans("ıİğĞüÜşŞöÖçÇ", "iigguussoocc")


def _pdf_paths() -> list[Path]:
    return sorted(PDF_DIR.glob("*.pdf"))


def _to_float(value: str) -> float:
    return float(value.replace(".", "").replace(",", "."))


def normalize_tanim(text: str) -> str:
    text = unicodedata.normalize("NFKD", text.translate(TR_MAP).lower())
    text = re.sub(r"[^\w\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _extract_tanim(parca: str) -> str:
    tanim = parca[:800]
    for num in FIYAT_RE.findall(tanim):
        tanim = tanim.replace(num, " ")
    tanim = BIRIM_RE.sub(" ", tanim)
    return re.sub(r"\s+", " ", tanim).strip()


def _ikili_fiyat_mi(path: Path, text: str) -> bool:
    name = path.name.upper()
    if "MEKAN" in name or "ELEK" in name:
        return True
    # Bazı listelerde birim + montaj/işçilik ayrı kolonda gelir.
    up = text.upper()
    return ("MONTAJ" in up and "BEDEL" in up) or ("ISCILIK" in up and "FIYAT" in up)


def _parse_pdf(path: Path) -> tuple[dict[str, float], dict[str, str]]:
    try:
        doc = fitz.open(path)
        text = "".join(page.get_text() for page in doc)
        doc.close()
        logger.debug(f"PDF açıldı: {path.name} ({len(text):,} karakter)")
    except Exception as e:
        logger.error(f"PDF okunamadı: {path} - {e}")
        return {}, {}
    
    iki_fiyat = _ikili_fiyat_mi(path, text)

    pozlar = POZ_RE.findall(text)
    parcalar = POZ_RE.split(text)
    fiyatlar: dict[str, float] = {}
    tanimlar: dict[str, str] = {}

    for poz, parca in zip(pozlar, parcalar[1:]):
        nums = FIYAT_RE.findall(parca[:800])
        if not nums:
            continue
        if iki_fiyat and len(nums) >= 2:
            fiyatlar[poz] = round(_to_float(nums[-2]) + _to_float(nums[-1]), 2)
        else:
            fiyatlar[poz] = _to_float(nums[-1])
        tanimlar[poz] = _extract_tanim(parca)

    logger.info(f"PDF işlendi: {path.name} - {len(fiyatlar)} poz bulundu")
    return fiyatlar, tanimlar


def _cache_guncel_mi() -> bool:
    if not CACHE_XLSX.is_file() or not CACHE_TANIM_XLSX.is_file():
        return False
    cache_mtime = min(CACHE_XLSX.stat().st_mtime, CACHE_TANIM_XLSX.stat().st_mtime)
    if not PDF_DIR.is_dir():
        return False
    for pdf in PDF_DIR.glob("*.pdf"):
        if pdf.stat().st_mtime > cache_mtime:
            return False
    return True


def _kaydet_cache(fiyatlar: dict[str, float], tanimlar: dict[str, str]) -> None:
    try:
        CACHE_XLSX.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CSB Fiyatlar"
        ws.append(["Poz No", "Birim Fiyat (TL)", "Kaynak"])
        for poz in sorted(fiyatlar):
            ws.append([poz, fiyatlar[poz], poz[:2]])
        wb.save(CACHE_XLSX)
        logger.info(f"Cache dosyası kaydedildi: {CACHE_XLSX}")
    except Exception as e:
        logger.error(f"Cache dosyası kaydedilemedi: {CACHE_XLSX} - {e}")

    try:
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "CSB Tanımlar"
        ws2.append(["Poz No", "Tanım"])
        for poz in sorted(tanimlar):
            ws2.append([poz, tanimlar[poz]])
        wb2.save(CACHE_TANIM_XLSX)
        logger.info(f"Tanım cache dosyası kaydedildi: {CACHE_TANIM_XLSX}")
    except Exception as e:
        logger.error(f"Tanım cache dosyası kaydedilemedi: {CACHE_TANIM_XLSX} - {e}")


def _yukle_cache() -> dict[str, float]:
    try:
        wb = openpyxl.load_workbook(CACHE_XLSX, data_only=True)
        ws = wb.active
        fiyatlar: dict[str, float] = {}
        for row in range(2, ws.max_row + 1):
            poz = ws.cell(row, 1).value
            fiyat = ws.cell(row, 2).value
            if poz is None or fiyat is None:
                continue
            try:
                fiyatlar[str(poz).strip()] = float(fiyat)
            except (TypeError, ValueError):
                continue
        wb.close()
        logger.info(f"Cache yüklendi: {len(fiyatlar)} poz")
        return fiyatlar
    except Exception as e:
        logger.error(f"Cache yüklenemedi: {CACHE_XLSX} - {e}")
        return {}


def _yukle_tanim_cache() -> dict[str, str]:
    try:
        wb = openpyxl.load_workbook(CACHE_TANIM_XLSX, data_only=True)
        ws = wb.active
        tanimlar: dict[str, str] = {}
        for row in range(2, ws.max_row + 1):
            poz = ws.cell(row, 1).value
            tanim = ws.cell(row, 2).value
            if poz and tanim:
                tanimlar[str(poz).strip()] = str(tanim)
        wb.close()
        logger.info(f"Tanım cache yüklendi: {len(tanimlar)} poz")
        return tanimlar
    except Exception as e:
        logger.error(f"Tanım cache yüklenemedi: {CACHE_TANIM_XLSX} - {e}")
        return {}


def pdflerden_fiyatlari_cikar(*, cache_kullan: bool = True) -> dict[str, float]:
    return pdflerden_verileri_cikar(cache_kullan=cache_kullan)[0]


def pdflerden_verileri_cikar(
    *, cache_kullan: bool = True
) -> tuple[dict[str, float], dict[str, str]]:
    if cache_kullan and _cache_guncel_mi():
        logger.debug("Güncel cache bulundu, yükleniyor")
        return _yukle_cache(), _yukle_tanim_cache()

    if not PDF_DIR.is_dir():
        logger.error(f"pdfler klasörü yok: {PDF_DIR}")
        raise FileNotFoundError(f"pdfler klasörü yok: {PDF_DIR}")

    try:
        pdfler = _pdf_paths()
        if not pdfler:
            logger.warning(f"pdfler klasöründe PDF dosyası bulunamadı: {PDF_DIR}")
        
        logger.info(f"PDF dosyalarını işleme başlanıyor: {len(pdfler)} dosya")
        fiyatlar: dict[str, float] = {}
        tanimlar: dict[str, str] = {}

        for pdf in pdfler:
            f, t = _parse_pdf(pdf)
            fiyatlar.update(f)
            tanimlar.update(t)

        logger.info(f"Toplam {len(fiyatlar)} poz işlendi")

        if cache_kullan:
            _kaydet_cache(fiyatlar, tanimlar)

        return fiyatlar, tanimlar
    except Exception as e:
        logger.error(f"PDF verileri çıkarılırken hata: {e}")
        raise


def _token_set(text: str) -> set[str]:
    return {w for w in normalize_tanim(text).split() if len(w) > 2}


def tanim_ile_eslestir(
    aranan: str,
    tanimlar: dict[str, str],
    fiyatlar: dict[str, float],
    *,
    min_skor: float = 0.52,
) -> tuple[str, str, float] | None:
    """Cetvel tanımına en yakın CSB pozunu döndürür: (poz, csb_tanim, skor)."""
    if not aranan or not aranan.strip():
        return None

    aranan_norm = normalize_tanim(aranan)
    aranan_tokens = _token_set(aranan)
    if not aranan_tokens:
        return None

    adaylar: list[tuple[float, str]] = []
    for poz, tanim in tanimlar.items():
        if poz not in fiyatlar:
            continue
        csb_tokens = _token_set(tanim)
        if not csb_tokens:
            continue
        ortak = len(aranan_tokens & csb_tokens)
        if ortak < 2 and ortak / len(aranan_tokens) < 0.25:
            continue
        on_skor = ortak / len(aranan_tokens)
        adaylar.append((on_skor, poz))

    adaylar.sort(reverse=True)
    en_iyi: tuple[str, str, float] | None = None
    for _, poz in adaylar[:40]:
        csb_tanim = tanimlar[poz]
        skor = SequenceMatcher(None, aranan_norm, normalize_tanim(csb_tanim)).ratio()
        if en_iyi is None or skor > en_iyi[2]:
            en_iyi = (poz, csb_tanim, skor)

    if en_iyi and en_iyi[2] >= min_skor:
        return en_iyi
    return None
