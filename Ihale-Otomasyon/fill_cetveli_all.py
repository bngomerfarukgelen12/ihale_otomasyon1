import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

# csb_poz_numaralari dosyasını oku
poz_df = pd.read_excel(r'c:\Users\argge\OneDrive\Masaüstü\Ihale-Otomasyon\csb_poz_numaralari.xlsx')
print(f"Poz tanımları yüklendi: {len(poz_df)} tanım")

# Belediye cetvelini aç
cetveli_path = r'c:\Users\argge\OneDrive\Masaüstü\Ihale-Otomasyon\excel\belediye teklif cetveli.xlsx'
wb = openpyxl.load_workbook(cetveli_path)
ws = wb.active

# Birleştirilmiş hücreleri bul ve kaldır
merged_ranges = list(ws.merged_cells.ranges)
print(f"Birleştirilmiş hücre aralıkları: {len(merged_ranges)}")
for merged_range in merged_ranges:
    print(f"  Unmerge: {merged_range}")
    ws.unmerge_cells(str(merged_range))

# Başlık satırları (1-5) korunacak
# Veri başlangıç satırı = 6

# Mevcut veri satırlarını temizle (6'dan 100'e)
for row in range(6, 100):
    for col in range(1, 9):  # A'dan H'ye
        ws.cell(row=row, column=col).value = None

# Yeni satırlar ekle (poz_df'deki 1878 satır için)
start_row = 6
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for idx, (_, row_data) in enumerate(poz_df.iterrows()):
    current_row = start_row + idx
    
    # Sıra No (A sütunu)
    ws[f'A{current_row}'].value = row_data['Sıra']
    ws[f'A{current_row}'].border = thin_border
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    
    # İş Kalemi No (B sütunu) - Poz No
    ws[f'B{current_row}'].value = row_data['Poz No']
    ws[f'B{current_row}'].border = thin_border
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
    
    # İş Kaleminin Adı (C sütunu) - Tanım
    ws[f'C{current_row}'].value = row_data['Tanım']
    ws[f'C{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws[f'C{current_row}'].border = thin_border
    
    # Birimi (D sütunu)
    ws[f'D{current_row}'].value = row_data['Birim']
    ws[f'D{current_row}'].border = thin_border
    ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
    
    if (idx + 1) % 200 == 0:
        print(f"  {idx + 1} satır işlendi...")

print(f"\nToplam {len(poz_df)} satır eklendi")

# Toplam Tutar formülünü ekle
last_row = start_row + len(poz_df) - 1
formula_row = last_row + 1

ws[f'A{formula_row}'].value = 'TOPLAM TUTAR (KDV Hariç)'
ws[f'A{formula_row}'].font = Font(bold=True)
ws[f'G{formula_row}'].value = f'=SUM(G{start_row}:G{last_row})'
ws[f'G{formula_row}'].font = Font(bold=True)

print(f"Toplam Tutar formülü: {ws[f'G{formula_row}'].value}")

# Dosyayı kaydet
wb.save(cetveli_path)
print(f"\nDosya kaydedildi: {cetveli_path}")
print(f"Toplam satır sayısı: {last_row - start_row + 1} (Sıra 1'den {len(poz_df)}'e)")
