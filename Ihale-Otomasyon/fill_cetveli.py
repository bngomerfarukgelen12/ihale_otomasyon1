import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

# Tüm tanımları oku
tanimlar_df = pd.read_excel(r'c:\Users\argge\OneDrive\Masaüstü\Ihale-Otomasyon\excel\csb_tum_tanimlar.xlsx')

# Poz No -> Tanım şeklinde dictionary yap
poz_dict = dict(zip(tanimlar_df['Poz No'], tanimlar_df['Tanım']))

print(f"Tanım sözlüğü oluşturuldu: {len(poz_dict)} tanım")
print(f"Örnek: {list(poz_dict.items())[:3]}")

# Belediye cetvelini aç
cetveli_path = r'c:\Users\argge\OneDrive\Masaüstü\Ihale-Otomasyon\excel\belediye teklif cetveli.xlsx'
wb = openpyxl.load_workbook(cetveli_path)
ws = wb.active

# Hangi poz numaraları kullanılıyor, kontrol et
print("\nBelediye cetvelinde bulunan İş Kalemi Numaraları:")
isi_kalemi_nos = {}
for row in range(6, 40):  # 6'dan 39'a kadar
    cell_value = ws[f'B{row}'].value
    sira_no = ws[f'A{row}'].value
    if cell_value is not None and cell_value != "İş Kalemi No":
        print(f"  Satır {row}: Sıra No = {sira_no}, Poz No = {cell_value}")
        isi_kalemi_nos[row] = cell_value

print(f"\nToplam {len(isi_kalemi_nos)} İş Kalemi bulundu")

# Şimdi tanımları ekle
satir_count = 0
for row, poz_no in isi_kalemi_nos.items():
    # Poz No ile tanımı bul
    tanım = poz_dict.get(poz_no, f"Tanım bulunamadı: {poz_no}")
    
    # C sütununa (İş Kaleminin Adı) tanımı yaz
    ws[f'C{row}'].value = tanım
    ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    print(f"Satır {row}: {poz_no} -> {tanım[:50]}...")
    satir_count += 1

print(f"\n{satir_count} satır güncellendi")

# Dosyayı kaydet
wb.save(cetveli_path)
print(f"\nDosya kaydedildi: {cetveli_path}")
