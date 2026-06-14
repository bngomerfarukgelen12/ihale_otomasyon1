#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ankara Belediyesi teklif cetvelindeki iş kalemlerinin birim fiyatlarını
CSB fiyat listesinden bulup dolduran script.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from difflib import SequenceMatcher
import sys
import io
import os

# Unicode çıktı için stdout'u ayarla
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def load_csb_prices():
    """CSB İnşaat Birim Fiyat listesini yükle"""
    excel_path = r'c:\Users\argge\OneDrive\Masaüstü\Ihale-Otomasyon\excel\csb insaat birim fiyatları.xlsx'
    
    # Row 1 (0-indexed) as header, skip row 0
    df = pd.read_excel(excel_path, sheet_name=0, header=1, skiprows=0)
    
    # Kolon adlarını normalize et (trailing space kaldır)
    df.columns = [str(col).strip() for col in df.columns]
    
    # Gerçek veri satırlarını filtrele (NaN olmayan Poz No)
    df = df.dropna(subset=['Poz No']).copy()
    df['Poz_No'] = df['Poz No'].astype(str).str.strip()
    
    # Birim fiyatı sayıya dönüştür
    df['Fiyat_num'] = pd.to_numeric(df['Fiyat'], errors='coerce')
    df = df[df['Fiyat_num'].notna()].copy()
    
    # Tanım ve Birim temizle
    df['Tanım'] = df['Tanım'].astype(str).str.strip()
    df['Birim'] = df['Birim'].astype(str).str.strip()
    
    print(f"OK: {len(df)} unit prices loaded from CSB price list")
    if len(df) > 0:
        print(f"  Example: {df.iloc[0]['Poz_No']} - {df.iloc[0]['Tanım'][:50]}... ({df.iloc[0]['Fiyat_num']} {df.iloc[0]['Birim']})")
    
    return df

def similarity_ratio(a, b):
    """İki string arasındaki benzerlik oranını hesapla"""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def find_best_match(tanım, birim, csb_df):
    """
    Ankara cetvelindeki tanımı CSB listesinde bul
    Benzerlik oranına göre en iyi eşleşmeyi döndür
    """
    best_match = None
    best_score = 0.0
    
    for idx, row in csb_df.iterrows():
        csb_tanım = row['Tanım']
        csb_birim = row['Birim']
        
        # Aynı birim ve yüksek benzerlik ara
        if birim.lower() in csb_birim.lower() or csb_birim.lower() in birim.lower():
            sim = similarity_ratio(tanım, csb_tanım)
            
            if sim > best_score:
                best_score = sim
                best_match = {
                    'poz_no': row['Poz_No'],
                    'tanım': csb_tanım,
                    'birim': csb_birim,
                    'fiyat': row['Fiyat_num'],
                    'similarity': sim
                }
    
    # Hiç birim eşleşmezse, sadece benzerliğe göre ara
    if best_match is None or best_score < 0.5:
        for idx, row in csb_df.iterrows():
            sim = similarity_ratio(tanım, row['Tanım'])
            
            if sim > best_score:
                best_score = sim
                best_match = {
                    'poz_no': row['Poz_No'],
                    'tanım': row['Tanım'],
                    'birim': row['Birim'],
                    'fiyat': row['Fiyat_num'],
                    'similarity': sim
                }
    
    return best_match if best_score > 0.4 else None

def main():
    """Ana fonksiyon"""
    print("=" * 80)
    print("Fill Ankara Municipality Tender List with CSB Unit Prices")
    print("=" * 80)
    print()
    
    # CSB fiyat listesini yükle
    csb_df = load_csb_prices()
    print()
    
    # Ankara Belediyesi cetvelini aç
    cetveli_path = r'c:\Users\argge\OneDrive\Masaüstü\Ihale-Otomasyon\excel\ANKARA B.BEL. TEKLİF CETVELİ .xlsx'
    wb = openpyxl.load_workbook(cetveli_path)
    ws = wb.active
    
    # İş kalemlerini oku ve birim fiyatları bul
    print("Searching for unit prices...\n")
    
    matches_found = 0
    matches_not_found = []
    
    # Row 6'dan başlıyor (header Row 5)
    for row in range(6, 100):  # Maksimum 100 satır
        sira_no = ws[f'A{row}'].value
        if sira_no is None:
            break
        
        tanım = ws[f'C{row}'].value
        birim = ws[f'D{row}'].value
        
        if tanım is None or birim is None:
            continue
        
        tanım = str(tanım).strip()
        birim = str(birim).strip()
        
        # Birim fiyat sütunu (F)
        fiyat_cell = ws[f'F{row}']
        
        if fiyat_cell.value is not None:
            # Zaten doldurulmuş
            continue
        
        # En iyi eşleşmeyi bul
        best_match = find_best_match(tanım, birim, csb_df)
        
        if best_match:
            fiyat_cell.value = best_match['fiyat']
            fiyat_cell.number_format = '#,##0.00'
            matches_found += 1
            
            sim_percent = best_match['similarity'] * 100
            print(f"[Row {row}] Price: {best_match['fiyat']:.2f} {best_match['birim']}")
            print(f"  Item: {tanım[:60]}...")
            print(f"  Match ({sim_percent:.0f}%): {best_match['tanım'][:60]}...")
            print(f"  Code: {best_match['poz_no']}")
            print()
        else:
            matches_not_found.append({
                'row': row,
                'tanım': tanım,
                'birim': birim
            })
            print(f"[Row {row}] NO MATCH FOUND")
            print(f"  Item: {tanım[:60]}...")
            print()
    
    # Sonuç raporu
    print("=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Successful matches: {matches_found}")
    print(f"Items without matches: {len(matches_not_found)}")
    
    if matches_not_found:
        print("\nItems without matches:")
        for item in matches_not_found:
            print(f"  - Row {item['row']}: {item['tanım'][:70]}")
    
    # Dosyayı kaydet
    wb.save(cetveli_path)
    print(f"\nFile saved: {cetveli_path}")
    print()
    
    return 0

if __name__ == '__main__':
    sys.exit(main())
