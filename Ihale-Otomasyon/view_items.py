import openpyxl

cetveli_path = r'c:\Users\argge\OneDrive\Masaüstü\Ihale-Otomasyon\excel\ANKARA B.BEL. TEKLİF CETVELİ .xlsx'
wb = openpyxl.load_workbook(cetveli_path)
ws = wb.active

print('Eşleşme Olmayan Kalemler:')
print('='*80)

# Row 23
print(f'\nRow 23:')
print(f'  Tanım: {ws["C23"].value}')
print(f'  Birim: {ws["D23"].value}')
print(f'  Miktar: {ws["E23"].value}')

# Row 26
print(f'\nRow 26:')
print(f'  Tanım: {ws["C26"].value}')
print(f'  Birim: {ws["D26"].value}')
print(f'  Miktar: {ws["E26"].value}')
