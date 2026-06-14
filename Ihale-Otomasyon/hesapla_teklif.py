"""Teklif cetvelindeki pozları CSB birim fiyatlarıyla eşleştirip tutar hesaplar."""
from __future__ import annotations

import logging
import re
import sys
from pathlib import Path

import openpyxl

from csb_pdf_fiyatlar import pdflerden_fiyatlari_cikar, pdflerden_verileri_cikar, tanim_ile_eslestir
from logger_config import setup_logger, log_program_start, log_program_end, log_summary

logger = setup_logger(__name__)

ROOT = Path(__file__).parent
EXCEL_DIR = ROOT / "excel"
POZ_RE = re.compile(r"^\d{2}\.\d{3}\.\d{4}$")
OZEL_POZ_RE = re.compile(r"^ÖZEL", re.I)

COL_SIRA = 1
COL_POZ = 2
COL_TANIM = 3
COL_MIKTAR = 5
COL_BIRIM_FIYAT = 6
COL_TUTAR = 7

CETVELLER = {
    "belediye": {"pattern": "belediye*.xlsx", "data_start_row": 7, "eslesme": "poz_once"},
    "tpao": {"pattern": "tpao*.xlsx", "data_start_row": 6, "eslesme": "poz_once"},
    "belpark": {
        "pattern": "*BEL PARK*.xlsx",
        "data_start_row": 6,
        "eslesme": "tanim_once",
        "min_tanim_skor": 0.45,
    },
    "seydisehir": {
        "pattern": "*BEL PARK*.xlsx",
        "data_start_row": 6,
        "eslesme": "tanim_once",
        "min_tanim_skor": 0.35,
    },
}


def normalize_poz(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or OZEL_POZ_RE.match(text):
        return None
    if POZ_RE.match(text):
        return text
    digits = re.sub(r"\D", "", text)
    if len(digits) == 9:
        return f"{digits[0:2]}.{digits[2:5]}.{digits[5:9]}"
    return None


def find_file(pattern: str) -> Path:
    matches = sorted(EXCEL_DIR.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"{pattern} bulunamadı: {EXCEL_DIR}")
    return matches[0]


def load_csb_prices(csb_path: Path | None = None) -> dict[str, float]:
    """Önce pdfler/, yoksa excel/csb*.xlsx kaynağından fiyat yükler."""
    try:
        prices = pdflerden_fiyatlari_cikar()
        logger.info(f"CSB fiyatları PDF'den yüklendi: {len(prices)} poz")
        return prices
    except FileNotFoundError as e:
        logger.warning(f"PDF kaynağı bulunamadı, Excel'den yükleme denenecek: {e}")

    try:
        if csb_path is None:
            csb_path = find_file("csb*.xlsx")
        
        logger.info(f"CSB fiyatları Excel'den yükleniyor: {csb_path}")
        wb = openpyxl.load_workbook(csb_path, data_only=True)
        ws = wb.active
        prices: dict[str, float] = {}
        for row in range(3, ws.max_row + 1):
            poz = ws.cell(row, 1).value
            fiyat = ws.cell(row, 4).value
            if poz is None or fiyat is None:
                continue
            key = str(poz).strip()
            if not POZ_RE.match(key):
                continue
            try:
                prices[key] = float(fiyat)
            except (TypeError, ValueError):
                continue
        wb.close()
        logger.info(f"CSB fiyatları Excel'den yüklendi: {len(prices)} poz")
        return prices
    except Exception as e:
        logger.error(f"CSB fiyatları yüklenemedi: {e}")
        raise


def find_total_row(ws, data_start_row: int) -> int:
    for row in range(data_start_row, ws.max_row + 1):
        val = ws.cell(row, COL_SIRA).value
        if isinstance(val, str) and "TOPLAM" in val.upper():
            return row
    return ws.max_row + 1


def hesapla_teklif(
    cetvel: str = "belediye",
    teklif_path: Path | None = None,
    csb_path: Path | None = None,
    *,
    kaydet: bool = True,
) -> dict:
    try:
        if cetvel not in CETVELLER:
            msg = f"Bilinmeyen cetvel: {cetvel}. Seçenekler: {', '.join(CETVELLER)}"
            logger.error(msg)
            raise ValueError(msg)

        cfg = CETVELLER[cetvel]
        data_start_row = cfg["data_start_row"]
        eslesme_modu = cfg.get("eslesme", "poz_once")
        min_tanim_skor = cfg.get("min_tanim_skor", 0.52)
        teklif_path = teklif_path or find_file(cfg["pattern"])
        csb_path = csb_path or (find_file("csb*.xlsx") if list(EXCEL_DIR.glob("csb*.xlsx")) else None)
        
        logger.info(f"Cetvel işleniyor: {cetvel}")
        logger.info(f"Teklif dosyası: {teklif_path}")
        
        prices, tanimlar = pdflerden_verileri_cikar()

        wb = openpyxl.load_workbook(teklif_path)
        ws = wb.active
        total_row = find_total_row(ws, data_start_row)

        satirlar: list[dict] = []
        toplam = 0.0
        bulunamayan: list[str] = []
        tanim_eslesen = 0

        for row in range(data_start_row, total_row):
            poz_raw = ws.cell(row, COL_POZ).value
            tanim_raw = ws.cell(row, COL_TANIM).value
            miktar_raw = ws.cell(row, COL_MIKTAR).value
            if miktar_raw is None or not tanim_raw:
                continue

            try:
                miktar = float(miktar_raw)
            except (TypeError, ValueError):
                logger.warning(f"Satır {row}: Miktar değeri dönüştürülemedi: {miktar_raw}")
                continue

            poz = normalize_poz(poz_raw)
            birim_fiyat = None
            eslesme = "poz"
            skor = 1.0
            csb_tanim = ""

            if eslesme_modu == "tanim_once":
                eslesme_sonuc = tanim_ile_eslestir(
                    str(tanim_raw), tanimlar, prices, min_skor=min_tanim_skor
                )
                if eslesme_sonuc:
                    poz, csb_tanim, skor = eslesme_sonuc
                    birim_fiyat = prices[poz]
                    ws.cell(row, COL_POZ).value = poz
                    eslesme = "tanim"
                    tanim_eslesen += 1
                    logger.debug(f"Tanımla eşleşti: {poz} (skor: {skor:.2f})")
                elif poz and poz in prices:
                    birim_fiyat = prices[poz]
                    eslesme = "poz"
                    logger.debug(f"Poz numarasıyla eşleşti: {poz}")
            else:
                birim_fiyat = prices.get(poz) if poz else None
                if birim_fiyat is None:
                    eslesme_sonuc = tanim_ile_eslestir(
                        str(tanim_raw), tanimlar, prices, min_skor=min_tanim_skor
                    )
                    if eslesme_sonuc:
                        poz, csb_tanim, skor = eslesme_sonuc
                        birim_fiyat = prices[poz]
                        ws.cell(row, COL_POZ).value = poz
                        eslesme = "tanim"
                        tanim_eslesen += 1
                        logger.debug(f"Tanımla eşleşti: {poz} (skor: {skor:.2f})")

            if birim_fiyat is None:
                etiket = str(poz_raw or tanim_raw)[:60]
                bulunamayan.append(etiket)
                logger.warning(f"Fiyat bulunamadı: {etiket}")
                continue

            tutar = round(miktar * birim_fiyat, 2)
            ws.cell(row, COL_BIRIM_FIYAT).value = birim_fiyat
            ws.cell(row, COL_TUTAR).value = tutar
            toplam += tutar

            satirlar.append(
                {
                    "sira": ws.cell(row, COL_SIRA).value,
                    "poz": poz,
                    "tanim": str(tanim_raw)[:80],
                    "csb_tanim": (csb_tanim or "")[:80],
                    "miktar": miktar,
                    "birim_fiyat": birim_fiyat,
                    "tutar": tutar,
                    "eslesme": eslesme,
                    "skor": skor,
                }
            )

        ws.cell(total_row, COL_TUTAR).value = round(toplam, 2)

        if kaydet:
            wb.save(teklif_path)
            logger.info(f"Teklif dosyası kaydedildi: {teklif_path}")
        wb.close()

        result = {
            "cetvel": cetvel,
            "teklif_dosyasi": str(teklif_path),
            "csb_dosyasi": "pdfler/ (tum PDF listeleri)" if (ROOT / "pdfler").is_dir() else str(csb_path or ""),
            "satirlar": satirlar,
            "toplam": round(toplam, 2),
            "bulunamayan": bulunamayan,
            "hesaplanan_sayisi": len(satirlar),
            "bulunamayan_sayisi": len(bulunamayan),
            "tanim_eslesen_sayisi": tanim_eslesen,
        }
        
        logger.info(f"İşlem tamamlandı: {len(satirlar)} kalem hesaplandı, {len(bulunamayan)} kalem bulunamadı")
        return result
    except Exception as e:
        logger.error(f"Teklif hesaplanırken hata: {e}")
        raise


def yazdir_sonuc(sonuc: dict) -> None:
    print(f"Cetvel: {sonuc['cetvel']}")
    print(f"Güncellenen dosya: {sonuc['teklif_dosyasi']}")
    print(f"CSB kaynağı: {sonuc['csb_dosyasi']}")
    print(
        f"Hesaplanan: {sonuc['hesaplanan_sayisi']} kalem  |  "
        f"Tanımla eşleşen: {sonuc.get('tanim_eslesen_sayisi', 0)}  |  "
        f"Bulunamayan: {sonuc['bulunamayan_sayisi']} kalem\n"
    )

    for s in sonuc["satirlar"]:
        ek = f" [{s['eslesme']}, %{s['skor']*100:.0f}]" if s.get("eslesme") == "tanim" else ""
        print(
            f"  {s['poz']}{ek}  |  miktar: {s['miktar']:,.0f}  |  "
            f"birim: {s['birim_fiyat']:,.2f} TL  |  tutar: {s['tutar']:,.2f} TL"
        )
        if s.get("eslesme") == "tanim" and s.get("csb_tanim"):
            metin = s["csb_tanim"].encode("cp1254", errors="replace").decode("cp1254")
            print(f"      -> {metin}")

    if sonuc["bulunamayan"]:
        print("\nCSB listelerinde bulunamayan pozlar:")
        for poz in sonuc["bulunamayan"]:
            print(f"  - {poz}")

    print(f"\nTOPLAM (hesaplanan kalemler, KDV hariç): {sonuc['toplam']:,.2f} TL")


def main(argv: list[str] | None = None) -> int:
    try:
        log_program_start(logger)
        
        if not EXCEL_DIR.is_dir():
            logger.error(f"excel klasörü yok: {EXCEL_DIR}")
            print(f"[HATA] excel klasörü yok: {EXCEL_DIR}", file=sys.stderr)
            log_program_end(logger, 1)
            return 1

        args = (argv if argv is not None else sys.argv)[1:]
        cetvel = args[0] if args else "belediye"
        if cetvel not in CETVELLER:
            msg = f"Kullanım: python hesapla_teklif.py [{('|').join(CETVELLER)}]"
            logger.error(msg)
            print(msg, file=sys.stderr)
            log_program_end(logger, 1)
            return 1

        sonuc = hesapla_teklif(cetvel)
        yazdir_sonuc(sonuc)
        
        summary = {
            "Cetvel": sonuc['cetvel'],
            "Hesaplanan kalem": sonuc['hesaplanan_sayisi'],
            "Tanim eslesen": sonuc.get('tanim_eslesen_sayisi', 0),
            "Bulunamayan kalem": sonuc['bulunamayan_sayisi'],
            "Toplam tutar": f"{sonuc['toplam']:,.2f} TL",
            "Teklif dosyasi": sonuc['teklif_dosyasi'],
        }
        log_summary(logger, summary)
        log_program_end(logger, 0)
        return 0
    except FileNotFoundError as e:
        logger.error(f"Dosya bulunamadı: {e}")
        print(f"[HATA] Dosya bulunamadı: {e}", file=sys.stderr)
        log_program_end(logger, 1)
        return 1
    except ValueError as e:
        logger.error(f"Deger hatasi: {e}")
        print(f"[HATA] Hata: {e}", file=sys.stderr)
        log_program_end(logger, 1)
        return 1
    except Exception as e:
        logger.exception(f"Beklenmeyen hata: {e}")
        print(f"[HATA] Beklenmeyen hata: {e}", file=sys.stderr)
        log_program_end(logger, 1)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
