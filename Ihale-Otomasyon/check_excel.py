import openpyxl

wb = openpyxl.load_workbook('excel/OKUL PENCERE KAPI.xlsx')
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print("\nFirst 30 rows:")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(row)
        if i >= 30:
            break
